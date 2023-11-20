from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Font
import openpyxl as pxl
import warnings
warnings.simplefilter('ignore')

workbook_1 = pxl.load_workbook(r'office1.xlsx')#WTS-005-001_Raw导出数据集_20230615_迁移前.xlsx
workbook_2 = pxl.load_workbook(r'wps2.xlsx')

workbook_1_sheet_names = workbook_1.sheetnames
workbook_2_sheet_names = workbook_2.sheetnames

a = len(workbook_1_sheet_names)

for i in range(0,a):
    workbook_1_sheet_1 = workbook_1[workbook_1_sheet_names[i]]
    workbook_2_sheet_1 = workbook_2[workbook_2_sheet_names[i]]

    if(workbook_1_sheet_names[i] != workbook_2_sheet_names[i]):
        print(workbook_1_sheet_names[i])
    else:
        print(workbook_1_sheet_names[i] +" "+workbook_2_sheet_names[i])

    max_row = workbook_1_sheet_1.max_row if workbook_1_sheet_1.max_row > workbook_2_sheet_1.max_row else workbook_2_sheet_1.max_row
    max_column = workbook_1_sheet_1.max_column if workbook_1_sheet_1.max_column > workbook_2_sheet_1.max_column else workbook_2_sheet_1.max_column

    for i in range(1, (max_row + 1)):

        for j in range(1, (max_column + 1)):

            cell_1 = workbook_1_sheet_1.cell(i, j)

            cell_2 = workbook_2_sheet_1.cell(i, j)
            #print("success")
            if cell_1.value != cell_2.value:
                cell_1.fill = PatternFill("solid", fgColor='FFFF00')

                cell_1.font = Font(color=colors.BLACK, bold=True)

                cell_2.fill = PatternFill("solid", fgColor='FFFF00')

                cell_2.font = Font(color=colors.BLACK, bold=True)
                print(cell_2.value)





workbook_1.save('data615-前.xlsx')
workbook_2.save('data615-后.xlsx')
