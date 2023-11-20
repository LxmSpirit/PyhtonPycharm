from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Font
import openpyxl as pxl
import warnings
warnings.simplefilter('ignore')

workbook_1 = pxl.load_workbook(r'data3.xlsx')

workbook_1_sheet_names = workbook_1.sheetnames

a = len(workbook_1_sheet_names)
yellow = []
yellowsheet=[]

for i in range(0,a):
    workbook_1_sheet_1 = workbook_1[workbook_1_sheet_names[i]]

    max_row = workbook_1_sheet_1.max_row
    max_column = workbook_1_sheet_1.max_column

    for i in range(1, (max_row + 1)):

        for j in range(1, (max_column + 1)):

            cell_1 = workbook_1_sheet_1.cell(i, j)
            #print(type(cell_1.value))
            #print(cell_1.value)
            if cell_1.fill.start_color.rgb == "00FFFF00" :
                yellow.append(cell_1.value)
                yellowsheet.append(workbook_1_sheet_1)
                #print(cell_1.value)


print(yellow)
print(yellowsheet)

#workbook_1.save('data5.xlsx')
