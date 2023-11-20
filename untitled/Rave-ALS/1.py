#导入数值计算库
import numpy as np
#导入科学计算库
import pandas as pd
#导入交叉验证库
#from sklearn import cross_validation
#导入GaussianNB库
#from sklearn.naive_bayes import GaussianNB
import xlrd
import os
import re
import jieba
import pandas as pd
import jieba
import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.naive_bayes import MultinomialNB
from sklearn.linear_model import LogisticRegression
from pandas import DataFrame
import openpyxl as op

URL = 'D:/Work/Python/PycharmProjects/untitled/Rave-ALS/ALS/95.xls'
formaddress = 'D:/Work/Python/PycharmProjects/untitled/Rave-ALS/formdata.txt'
fieldaddress = 'D:/Work/Python/PycharmProjects/untitled/Rave-ALS/fieldata.txt'
def readforms():
    workbook = xlrd.open_workbook(URL)
    #print(workbook.sheet_names())
    sheet1 = workbook.sheet_by_name('Forms')
    nrows = sheet1.nrows #行数
    ncols = sheet1.ncols #列数
    dicform = {}
    for i in range(1,nrows):
        cellOID = sheet1.cell(i,0).value
        cellName = sheet1.cell(i, 2).value
        newOID = re.sub(r'\d+', '', cellOID)
        pattern = re.compile('[!@#$%^&*\(\)（）_+[\]{};:,./<>?\|`~-]')
        cellName2 = re.sub(pattern, '', cellName)
        newName = jieba.lcut_for_search(cellName2)
        s = ""
        for name in newName:
            s = s + " " + name

        if newOID in dicform:
            oldv = dicform[newOID]
            dicform[newOID] = oldv+" "+ s
        else:
            dicform[newOID] =  s
        #print(cellOID, newOID,cellName,s)
        #print( newOID, s)
    #print(nrows,ncols)
    #print(dicform)
    return dicform

def readfields():
    workbook = xlrd.open_workbook(URL)
    #print(workbook.sheet_names())
    sheet1 = workbook.sheet_by_name('Fields')
    nrows = sheet1.nrows #行数
    ncols = sheet1.ncols #列数
    dicfield = {}
    #print(nrows, ncols)
    for i in range(1,nrows):
        formOID = sheet1.cell(i,0).value
        fieldOID = sheet1.cell(i,1).value
        fieldName2 = sheet1.cell(i, 14).value
        pattern = re.compile('[!？@#$%^&*\(\)（）_+[\]{};:,./<>?\|`~-]')
        fieldName = re.sub(pattern, '', fieldName2)
        if (formOID in fieldOID):
            newfieldOID = fieldOID.replace(formOID,"")
        else :
            newfieldOID = fieldOID

        newName = jieba.lcut_for_search(fieldName)
        s = ""
        for name in newName:
            s = s + " " + name
        if newfieldOID in dicfield:
            oldv = dicfield[newfieldOID]
            dicfield[newfieldOID] = oldv+" "+ s
        else:
            dicfield[newfieldOID] = s
        #print(formOID,fieldOID ,newfieldOID, fieldName)
        #print(nrows, ncols)
        #dicfield[newfieldOID] = fieldName
    return dicfield


def write(data,address):
    file = open(address,mode='w',encoding='utf-8')
    for k,v in data.items():
        file.write(str(k)+'_'+str(v)+'\n')



if __name__ == '__main__':
    dicform = readforms()
    #write(dicform,formaddress)
    dicfield = readfields()
    #write(dicfield,fieldaddress)

    for k,v in dicform.items():
        words = str(v).split()
        print(words)
        s = pd.Series(words)
        #cv = CountVectorizer(ngram_range=(1,2))
        #a = cv.fit(words)
        print( s )



