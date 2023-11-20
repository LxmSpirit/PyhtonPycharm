import openpyxl
import warnings
warnings.simplefilter('ignore')

filename="data3.xlsx"#读取excel
workbook=openpyxl.load_workbook(filename)


workbook_sheet_names = workbook.sheetnames

a = len(workbook_sheet_names)
yellow = []


for i in range(0,a):



    worksheet = workbook.get_sheet_by_name(workbook_sheet_names[i])  # 读取sheet
    rows, cols = worksheet.max_row, worksheet.max_column
    for i in range(1, (rows+1)):
        for j in range(1, (cols+1)):
            ce = worksheet.cell(row=i, column=j)
            fill = ce.fill
            if fill.start_color.rgb == "FFFF00" :
                yellow.append(ce.value)

print(yellow)


